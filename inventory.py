from openpyxl import Workbook, load_workbook
from openpyxl.styles import *


import warnings
warnings.filterwarnings('ignore')

'''
需要三张表
sections - 款号清单
products - 单品清单
erp - 实际库存
'''

wb1 = load_workbook('products.xlsx')
wb2 = load_workbook('sections.xlsx')
wb3 = load_workbook('erp.xlsx')

ws1 = wb1['Sheet0']
ws2 = wb2['Sheet1']
ws3 = wb3['download']

def product_search():
    i1 = 2
    invent = []
    while(True):
        code = ws1.cell(i1,3).value
        if code == None:
            break

        number = section_search(code)
        name = ws1.cell(i1, 2).value
        invent.append([name, code, number])
        i1 = i1 +1
        # print('prodect get:')
        # print(number)
    return(invent)


def section_search(code):
    i2 = 2

    while(True):
        section = ws2.cell(i2,1).value
        if section == None:
            break
        for x in range(3,7):
            section = ws2.cell(i2, x).value
            if section == code:
                number = 0
                for x in range(3,7):
                    check = ws2.cell(i2, x).value
                    if check != None:
                        x = erp_search(check)
                        number = number + x
                # print('section return:')
                # print(number)
                return(number)

        i2 = i2 + 1


def erp_search(check):
    i3 = 2
    number = 0
    while(True):
        erp_section = ws3.cell(i3, 6).value
        if erp_section == None:
            break
        print('check is :'+check)
        print(erp_section)
        if erp_section == check:
            x = ws3.cell(i3, 7).value
            number = number + int(x)
            print(x)
        i3 = i3 + 1
        # print('erp return:')
        # print(number)
    return(number)


def new_invent(invent):
    wb = Workbook()
    ws = wb.create_sheet('sheet1')
    x = 0
    for i in invent:
        ws.cell(x+2, 1).value = invent[x][0]
        ws.cell(x+2, 2).value = invent[x][1]
        ws.cell(x+2, 3).value = invent[x][2]
        x = x + 1
    wb.save('new_invent.xlsx')
  


if __name__ == '__main__':
    invent = product_search()
    new_invent(invent)

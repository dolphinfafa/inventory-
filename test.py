from json import load
from openpyxl import Workbook, load_workbook
from openpyxl.styles import *


import warnings
warnings.filterwarnings('ignore')

'''
需要三张表
sections - 款号清单
products - 单品清单
erp - 实际库存
'''

# wb = load_workbook('库存.xlsx')
# ws = wb.active

# i = 3
# while(True):
#     name = ws.cell(i, 1).value
#     if name == None:
#         break
#     print(name)
#     i += 1

a = 'jjjj'
b = a + '.xlsx'
print(b)